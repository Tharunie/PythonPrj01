import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('hotelData.xlsx')
ds = wb['dataset']
cell=ds.cell(1, 27)
print(cell.value)
print(ds.max_row)

for row in range(2, ds.max_row + 1):
    cell = ds.cell(row, 27)
    corrected = cell.value * 0.9756
    print(corrected)
    corrected_cell=ds.cell(row, 27)
    corrected_cell.value = corrected


values = Reference(ds, min_row=2, max_row=ds.max_row, min_col=27, max_col= 27)
chart=BarChart()
chart.add_data(values)
ds.add_chart(chart, 'ah2')
wb.save('hotelData2.xlsx')
